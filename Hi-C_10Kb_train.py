import os.path as osp
import random

import torch
import pandas as pd
import numpy as np

import matplotlib.pyplot as plt

import torch.nn.functional as F
import torch_geometric.transforms as T
from torch_geometric.nn import GCNConv, GAE, ARMAConv
from torch_geometric.data import DataLoader
from torch.autograd import Variable
import os
import openpyxl
import math
from utils import *
from skimage.metrics import structural_similarity as ssim

KERNEL_SIZE = 10
EPOCH = 300
BATCH_SIZE = 16
LR = 0.0001
INPUT = 128


dev = torch.device("cuda")
torch.set_default_tensor_type(torch.FloatTensor)


def reset(nn):
    def _reset(item):
        if hasattr(item, 'reset_parameters'):
            item.reset_parameters()

    if nn is not None:
        if hasattr(nn, 'children') and len(list(nn.children())) > 0:
            for item in nn.children():
                _reset(item)
        else:
            _reset(nn)


class InnerProductDecoder(torch.nn.Module):
    """
    --------------------------------
    描述 :  GAE中的解码层，z乘z的转置
    --------------------------------
    """
    def __init__(self):
        super(InnerProductDecoder, self).__init__()
        self.align = torch.nn.Linear(INPUT, INPUT)

    def forward(self, z):
        adj = torch.matmul(self.align(z), z.t())
        return adj



class Encoder(torch.nn.Module):
    """
    --------------------------------
    描述 :    GAE中的编码层-使用ARMA实现
    --------------------------------
    """

    def __init__(self, in_channels, out_channels):
        super(Encoder, self).__init__()
        self.conv1 = ARMAConv(in_channels, 2 * out_channels, num_stacks=3, num_layers=2)
        self.conv2 = ARMAConv(2 * out_channels, out_channels, num_stacks=3, num_layers=2)

    def forward(self, x, edge_index, edge_weight):
        output = F.relu(self.conv1(x=x, edge_index=edge_index, edge_weight=edge_weight))
        return F.relu(self.conv2(x=output, edge_index=edge_index, edge_weight=edge_weight))


class dnn(torch.nn.Module):
    """
    --------------------------------
    描述 :   一维卷积自编码
    --------------------------------
    """

    def __init__(self):
        super(dnn, self).__init__()
        self.dropout = torch.nn.Dropout(p=0.5)
        self.encoder = torch.nn.Sequential(
            torch.nn.ConvTranspose1d(in_channels=INPUT, out_channels=4 * INPUT, kernel_size=(1,)),
            torch.nn.ReLU(),
            torch.nn.ConvTranspose1d(in_channels=4 * INPUT, out_channels=8 * INPUT, kernel_size=(1,)),
            torch.nn.ReLU(),
            torch.nn.ConvTranspose1d(in_channels=8 * INPUT, out_channels=16 * INPUT, kernel_size=(1,)),
            torch.nn.ReLU(),
            torch.nn.ConvTranspose1d(in_channels=16 * INPUT, out_channels=64 * INPUT, kernel_size=(1,)),
            torch.nn.ReLU()
        )

        self.decoder = torch.nn.Sequential(
            torch.nn.Conv1d(in_channels=64 * INPUT, out_channels=16 * INPUT, kernel_size=(1,)),
            torch.nn.ReLU(),
            torch.nn.Conv1d(in_channels=16 * INPUT, out_channels=8 * INPUT, kernel_size=(1,)),
            torch.nn.ReLU(),
            torch.nn.Conv1d(in_channels=8 * INPUT, out_channels=4 * INPUT, kernel_size=(1,)),
            torch.nn.ReLU(),
            torch.nn.Conv1d(in_channels=4 * INPUT, out_channels=INPUT, kernel_size=(1,)),
            torch.nn.ReLU()
        )

    def forward(self, input):
        # i = self.dropout(input)
        encode = self.encoder(input)
        # encode = self.dropout(encode)
        decode = self.decoder(encode)
        return decode


class GAE_DNN(GAE):
    """
    --------------------------------
    描述 :  GAE主要代码，初始化所有网络的参数
    --------------------------------
    """

    def __init__(self, encoder, decoder, dnn):
        super(GAE_DNN, self).__init__(encoder, decoder)
        self.dnner = dnn
        GAE_DNN.reset_parameters(self)

    def reset_parameters(self):
        reset(self.encoder)
        reset(self.decoder)
        reset(self.dnner)


def data_load():
    """
    --------------------------------
    描述 :    根据工具类utils构建自己的数据集（详见PYG文档的构建数据集）实现mini-batch
    --------------------------------
    """

    print("读取数据")

    data_len = feature.size(0)
    data_list = []
    for i in range(0, data_len):
        node_feature = feature[i]
        edge_index = torch.nonzero(hic[i]).long().t()
        edge_weight = hic[i][edge_index[0], edge_index[1]]
        # edge_weight = edge_weight.view(1, edge_weight.size(0))
        GCNData = MyData(x=torch.FloatTensor(node_feature.float()),
                         edge_index=torch.LongTensor(edge_index.long()),
                         y=torch.FloatTensor(sprite[i][KERNEL_SIZE:KERNEL_SIZE*2, 0:KERNEL_SIZE].float()),
                         weight=torch.FloatTensor(edge_weight.float())
                         )
        data_list.append(GCNData)
    # sample_num = len(data_list)
    random.shuffle(data_list)
    # test_num = int(TEST_RATIO * sample_num)
    # test_num = test_num if test_num >= 1 else 1
    # test = DataLoader(test_list, batch_size=32)
    # train = DataLoader(train_list, batch_size=32)

    return data_list


def sum_row_col(input_matrix):
    """
    --------------------------------
    描述 :  求行和列和拼接
    --------------------------------
    """

    return torch.cat((torch.sum(input_matrix, dim=0), torch.sum(input_matrix, dim=1)))


def pearson_calculate(x, y):
    """
    --------------------------------
    描述 :   根据行和列和求两个矩阵的相似性
    --------------------------------
    """

    a = sum_row_col(x).numpy()
    b = sum_row_col(y).numpy()
    c = np.corrcoef(a, b)[0][1]
    rangex = np.max(np.array(x))-np.min(np.array(x))
    rangey = np.max(np.array(y))-np.min(np.array(y))
    range = max(rangex, rangey)
    d = ssim(np.array(x), np.array(y), data_range=range)
    return c, d


def test_av_pearson():
    """
    --------------------------------
    描述 :   将训练样本和回归后的进行相似性计算
    --------------------------------
    """

    pearson = []
    struct = []
    for i in range(hic.size(0)):
        x = hic[i][KERNEL_SIZE:2 * KERNEL_SIZE, 0:KERNEL_SIZE]
        y = sprite[i][KERNEL_SIZE:2 * KERNEL_SIZE, 0:KERNEL_SIZE]
        # y = torch.matmul(test_y[i], test_y[i].t())[10:20, 0:10]
        # x = torch.matmul(test_x[i], test_x[i].t())[10:20, 0:10]
        # a = x.cpu().numpy().reshape(1,100)
        # b = y.cpu().numpy().reshape(1,100)
        c, d = pearson_calculate(x, y)
        if (not math.isnan(c)):
            pearson.append(c)
        if (not math.isnan(d)):
            struct.append(d)
    print(np.mean(pearson))
    print(np.mean(struct))
    worksheet.cell(2, 1, value=np.mean(pearson))
    worksheet.cell(2, 2, value=np.mean(struct))


def train(epoch):
    """
    --------------------------------
    描述 :   训练
    --------------------------------
    """

    model.train()

    all_loss = []
    # random.shuffle(data_list)
    for batch_idx, data in enumerate(train_data):
        # for data in data_list:
        data = data.to(dev)
        optimizer.zero_grad()

        gcn = model.encoder(data.x, data.edge_index, data.weight)
        dnn = model.dnner(gcn.view(1, gcn.size(0), INPUT).permute(0, 2, 1)).permute(0, 2, 1).view(gcn.size(0), INPUT)
        predict = model.decoder(dnn)
        pre_trans = torch.ones(data.y.size(0), KERNEL_SIZE, KERNEL_SIZE)
        for i in range(data.y.size(0)):
            tmp = predict[i * 2 * KERNEL_SIZE + KERNEL_SIZE:i * 2 * KERNEL_SIZE + 2 * KERNEL_SIZE,
                  i * 2 * KERNEL_SIZE:i * 2 * KERNEL_SIZE + KERNEL_SIZE]
            pre_trans[i] = tmp
        loss = loss_fn(pre_trans.to(dev), data.y)
        all_loss.append(loss.item())
        loss.backward()
        optimizer.step()
        data = data.cpu()
        data = data.y
    print('epoch: {}, loss: {:.8f}'.format(epoch, np.mean(all_loss)))
    return np.mean(all_loss)


def test():
    model.eval()
    with torch.no_grad():
        p_predict_y_list = []
        s_predict_y_list = []
        for data in data_list:
            data = data.to(dev)
            gcn = model.encoder(data.x, data.edge_index, data.weight)
            dnn = model.dnner(gcn.view(1, 2 * KERNEL_SIZE, INPUT).permute(0, 2, 1)).permute(0, 2, 1).view(
                2 * KERNEL_SIZE, INPUT)
            predict = model.decoder(dnn)
            p_predict_y, s_predict_y = pearson_calculate(predict[KERNEL_SIZE:KERNEL_SIZE*2, 0:KERNEL_SIZE].cpu(), data.y.cpu())
            if not math.isnan(p_predict_y):
                p_predict_y_list.append(p_predict_y)
            if not math.isnan(s_predict_y):
                s_predict_y_list.append(s_predict_y)
        p = np.mean(p_predict_y_list)
        s = np.mean(s_predict_y_list)
        print('predict_y_pearson: {:.5f}'.format(p))
        print('predict_y_struct: {:.5f}'.format(s))
        return p, s



for chr in range(1, 24):
    v = #设置序号
    sprite = torch.load("./output/version" + str(v) + "/chr" + str(chr) + "/x_sample_10kb.pt")
    hic = torch.load("./output/version" + str(v) + "/chr" + str(chr) + "/y_sample_10kb.pt")
    feature = torch.load("./output/version" + str(v) + "/chr" + str(chr) + "/feature_10kb.pt")

    data_list = data_load()
    train_data = DataLoader(data_list, batch_size=BATCH_SIZE, shuffle=True)
    model_dnn = dnn().to(dev)
    model_dnn = torch.nn.DataParallel(model_dnn, device_ids=[0, 1])
    model_gcn = Encoder(data_list[0].x.size(1), INPUT).to(dev)
    model_gcn = torch.nn.DataParallel(model_gcn, device_ids=[0, 1])
    model = GAE_DNN(encoder=model_gcn, decoder=InnerProductDecoder().to(dev), dnn=model_dnn).to(dev)
    optimizer = torch.optim.Adagrad([
        {'params': model.encoder.parameters()},
        {'params': model.dnner.parameters(), 'lr': 0.0001}], lr=LR)
    loss_fn = torch.nn.SmoothL1Loss()

    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet('My Worksheet')
    worksheet.cell(1, 1, value='pearson')
    worksheet.cell(1, 2, value='struct')
    worksheet.cell(1, 3, value='loss')

    test_av_pearson()

    print("开始回归")
    for epoch in range(EPOCH):
        loss = train(epoch)
        pearson = 0
        struct = 0
        if ((epoch + 1) % 10 == 0):
            pearson, struct = test()
        worksheet.cell(epoch + 3, 1, value=pearson)
        worksheet.cell(epoch + 3, 2, value=struct)
        worksheet.cell(epoch + 3, 3, value=loss)

    model.eval()
    workbook.save("./output/version" + str(v) + "/chr" + str(chr) + "/10kb_log.xlsx")
    print("chr" + str(chr))
    torch.save(model, "./output/version" + str(v) + "/chr" + str(chr) + "/epoch_300_conv1d_ARMA_10kb_out_128.pkl")
